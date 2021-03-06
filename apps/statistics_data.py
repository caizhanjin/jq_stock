"""
数据分析统计脚本
"""
from datetime import datetime

from apps.orm_sqlite import JqBarData, JqStockInfo, StatDailyData
from peewee import JOIN
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.utility import load_json, get_trade_days


def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet, engine="openpyxl").fillna('-')
        df.loc[len(df)] = list(df.columns)
        for col in df.columns:
            index = list(df.columns).index(col)
            letter = get_column_letter(index+1)
            collen = df[col].apply(lambda x: len(str(x).encode())).max()
            ws.column_dimensions[letter].width = min(collen*1.2, 18)

    wb.save(filename)


def stat_script_main(stat_date):
    print(f"\n开始统计 {stat_date} 数据...")

    bar_list = (
        JqBarData.select(JqBarData, JqStockInfo)
        .join(
            JqStockInfo,
            on=(JqBarData.index == JqStockInfo.index),
            join_type=JOIN.LEFT_OUTER
        )
        .where((JqBarData.datetime == stat_date) & (JqStockInfo.type == "stock"))
        .order_by(JqBarData.money.desc())
    )
    bar_dict = []

    for bar in bar_list:
        # 涨幅
        if bar.close and bar.pre_close:
            change_percent = round((bar.close - bar.pre_close) / bar.pre_close * 100, 2)
        else:
            change_percent = 0

        bar_dict.append({
            "code": bar.index,
            "display_name": "" if not bar.jqstockinfo else bar.jqstockinfo.display_name,
            "change_percent": change_percent,
            "money": 0 if not bar.money else bar.money,
            # "datetime": bar.datetime.strftime("%Y-%m-%d %H:%M:%S"),
            # "interval": bar.interval,

            "open": bar.open,
            "close": bar.close,
            "low": bar.low,
            "high": bar.high,
            "volume": bar.volume,
            "pre_close": bar.pre_close,

            "zjw_name": "" if not bar.jqstockinfo else bar.jqstockinfo.zjw_name,
            "sw_l1_name": "" if not bar.jqstockinfo else bar.jqstockinfo.sw_l1_name,
            "sw_l2_name": "" if not bar.jqstockinfo else bar.jqstockinfo.sw_l2_name,
            "sw_l3_name": "" if not bar.jqstockinfo else bar.jqstockinfo.sw_l3_name,

            "is_etf50": "否" if not bar.jqstockinfo else ("是" if bar.jqstockinfo.is_etf50 == 1 else "否"),
            "is_if300": "否" if not bar.jqstockinfo else ("是" if bar.jqstockinfo.is_if300 == 1 else "否"),
            "is_csi500": "否" if not bar.jqstockinfo else ("是" if bar.jqstockinfo.is_csi500 == 1 else "否"),
            "is_science": "否" if not bar.jqstockinfo else ("是" if bar.jqstockinfo.is_science == 1 else "否"),
        })

    df = pd.DataFrame.from_dict(bar_dict, orient="columns")
    df.set_index(["code"], inplace=True)
    df.sort_values(by=["money"], ascending=[False], inplace=True)

    # excel_writer初始化
    excel_name = f"{stat_date}_" \
                 f"{df[(df['money'] >= 2000000000.0)].shape[0]}_" \
                 f"{df[(df['money'] >= 200000000.0)].shape[0]}_" \
                 f"{df[(df['money'] >= 100000000.0)].shape[0]}_" \
                 f"{int(df[(df['is_etf50'] == '是')]['money'].sum())}_" \
                 f"{int(df[(df['is_if300'] == '是')]['money'].sum())}_" \
                 f"{int(df[(df['is_csi500'] == '是')]['money'].sum())}_" \
                 f"{int(df[(df['is_science'] == '是')]['money'].sum())}" \
                 f".xlsx"
    excel_name_path = os.path.join("output_data", excel_name)
    excel_writer = pd.ExcelWriter(excel_name_path)

    # daily_data数据查询和导出
    queryset = (
        StatDailyData.select()
        .where((StatDailyData.stat_date <= stat_date))
        .order_by(StatDailyData.stat_date.desc())
        .limit(60)
    )
    daily_list = [
        {
            "stat_date": i.stat_date,

            "science_money": i.science_money,
            "etf50_money": i.etf50_money,
            "if300_money": i.if300_money,
            "csi500_money": i.csi500_money,

            "front10_money": i.front10_money,
            "front15_money": i.front15_money,
            "front20_money": i.front20_money,
        }
        for i in queryset
    ]
    df_daily = pd.DataFrame.from_dict(daily_list, orient="columns")
    df_daily.set_index(["stat_date"], inplace=True)
    df_daily.to_excel(excel_writer, "daily_data")

    df.to_excel(excel_writer, "All")  # 所有数据写入
    df.sort_values(["sw_l1_name", "sw_l2_name", "sw_l3_name"], ascending=[1, 1, 1], inplace=True)

    all_stat_dict = {}
    # 成交额统计1
    df_2 = df[(df['money'] >= 200000000.0)]
    df_1 = df[(df['money'] >= 100000000.0)]
    print(f"\n当日成交额大于2亿：{df_2.shape[0]}，"
          f"\n成交金额大于1亿：{df_1.shape[0]}")
    df_1.to_excel(excel_writer, f"1亿以上")
    df_2.to_excel(excel_writer, f"2亿以上")

    df_20 = df[df["money"] >= 2000000000.0]
    print(f"\n成交额20亿以上："
          f"上涨：{df_20[(df_20['change_percent'] > 0)].shape[0]}，"
          f"下跌：{df_20[(df_20['change_percent'] < 0)].shape[0]}，"
          f"平：{df_20[(df_20['change_percent'] == 0)].shape[0]}")
    df_20.to_excel(excel_writer, f"20亿以上")

    # 成交前100、200涨跌家数和区间统计
    change_dict = {
        "前100": {},
        "前200": {},
    }
    df_dict = {
        "前100": df[:100],
        "前200": df[:200],
    }

    for key, item in change_dict.items():
        change_df = df_dict[key]
        change_dict[key] = {
            "overall": {
                "count": change_df.shape[0],
                "up": change_df[(change_df["change_percent"] > 0)].shape[0],
                "down": change_df[(change_df["change_percent"] < 0)].shape[0],
                "stay": change_df[(change_df["change_percent"] == 0)].shape[0],
            },
            "info": {
                "涨停": change_df[(change_df["change_percent"] >= 9.9)].shape[0],
                "7~9.9%": change_df[(change_df["change_percent"] >= 7) & (change_df["change_percent"] < 9.9)].shape[0],
                "5~7%": change_df[(change_df["change_percent"] >= 5) & (change_df["change_percent"] < 7)].shape[0],
                "2~5%": change_df[(change_df["change_percent"] >= 2) & (change_df["change_percent"] < 5)].shape[0],
                "0~2%": change_df[(change_df["change_percent"] > 0) & (change_df["change_percent"] < 2)].shape[0],
                "平": change_df[(change_df["change_percent"] == 0)].shape[0],
                "-0~2%": change_df[(change_df["change_percent"] > -2) & (change_df["change_percent"] < 0)].shape[0],
                "-2~5%": change_df[(change_df["change_percent"] > -5) & (change_df["change_percent"] <= -2)].shape[0],
                "-5~7%": change_df[(change_df["change_percent"] > -7) & (change_df["change_percent"] <= -5)].shape[0],
                "-7~9.9%": change_df[(change_df["change_percent"] > -9.9) & (change_df["change_percent"] <= -7)].shape[0],
                "跌停": change_df[(change_df["change_percent"] <= -9.9)].shape[0],
            }
        }
        print(f"\n{key}涨跌家数和区间统计>>\n"
              f"总览：{change_dict[key]['overall']}\n"
              f"详情：{change_dict[key]['info']}")

    # 成交额统计2
    # print(f"\n成交金额统计："
    #       f"前10总成交金额：{df[:10]['money'].sum()}，"
    #       f"前15总成交金额：{df[:15]['money'].sum()}，"
    #       f"前20总成交金额：{df[:20]['money'].sum()}，")

    # 00000000.0
    money_stat = {}
    _df = df[(df["money"] >= 10000000000.0)]
    money_stat["100亿以上"] = _df.shape[0]
    _df.to_excel(excel_writer, f"100亿以上")

    _df = df[(df["money"] >= 8000000000.0) & (df["money"] < 10000000000.0)]
    money_stat["80(含)~100亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"80(含)~100亿")

    _df = df[(df["money"] >= 5000000000.0) & (df["money"] < 8000000000.0)]
    money_stat["50(含)~80亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"50(含)~80亿")

    _df = df[(df["money"] >= 5000000000.0) & (df["money"] < 8000000000.0)]
    money_stat["50(含)~80亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"50(含)~80亿")

    _df = df[(df["money"] >= 4000000000.0) & (df["money"] < 5000000000.0)]
    money_stat["40(含)~50亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"40(含)~50亿")

    _df = df[(df["money"] >= 3000000000.0) & (df["money"] < 4000000000.0)]
    money_stat["30(含)~40亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"30(含)~40亿")

    _df = df[(df["money"] >= 2000000000.0) & (df["money"] < 3000000000.0)]
    money_stat["20(含)~30亿"] = _df.shape[0]
    _df.to_excel(excel_writer, f"20(含)~30亿")

    print(f"\n成交额区间划分：{money_stat}")

    excel_writer.save()
    reset_col(excel_name_path)
    pass


def main():
    # 统计脚本（主）
    """
    """
    stat_script_main(stat_date="2020-12-07")


if __name__ == '__main__':
    # reset_col(r"output_data\2020-12-07_43_879_1586.xlsx")
    # stat_script_main(stat_date="2020-12-07")

    main()

    pass
